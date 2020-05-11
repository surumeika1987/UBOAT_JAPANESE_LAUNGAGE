import openpyxl

JAPANESE_ROW = 6

if __name__ == "__main__":
    original_file = openpyxl.load_workbook('data\\Locales.xlsx')
    trans_file = openpyxl.load_workbook('data\\J_Locales.xlsx')
    additional_trans_file = openpyxl.load_workbook('data\\Addition_Locales.xlsx')

    original_sheet_names = original_file.sheetnames
    trans_sheet_names = trans_file.sheetnames

    trans_dict = {}
    missing_dict = {}

    # Load Keys
    for sheet_name in original_sheet_names:
        print("Loading Original Sheet: " + sheet_name)
        sheet = original_file[sheet_name]

        for i in range(sheet.max_row):
            cell = sheet.cell(row=i+1, column=1)
            if cell.value is not None:
                if not cell.value.startswith("/") and not cell.value.startswith('Key'):
                    trans_dict[(sheet_name, cell.value)] = ""
                    missing_dict[(sheet_name, cell.value)] = sheet.cell(row=i+1, column=2).value

    # Load Trans
    for sheet_name in trans_sheet_names:
        print("Loading Trans Sheet: " + sheet_name)
        sheet = trans_file[sheet_name]

        for i in range(sheet.max_row):
            key_cell = sheet.cell(row=i+1, column=1)
            if key_cell.value is not None:
                if (sheet_name, key_cell.value) in trans_dict:
                    if sheet.cell(row=i+1, column=JAPANESE_ROW).value is not None \
                            and 0 < len(str(sheet.cell(row=i+1, column=JAPANESE_ROW).value)):
                        trans_dict[(sheet_name, key_cell.value)] = sheet.cell(row=i+1, column=JAPANESE_ROW).value

    # Load Additional Trans
    for sheet_name in additional_trans_file.sheetnames:
        sheet = additional_trans_file[sheet_name]
        for i in range(sheet.max_row):
            key_cell = sheet.cell(row=i+1, column=1)
            if key_cell.value is not None:
                sheet_key, value_key = key_cell.value.split("->")
                trans_dict[(sheet_key, value_key)] = sheet.cell(row=i+1, column=2).value

    # Check Trans Data
    new_dict = {}
    for key in trans_dict:
        trans = trans_dict[key]
        trans = str(trans)
        if trans is None:
            print("Can't Found Trans Data: " + key[0] + "." + key[1])
        elif len(trans) == 0:
            print("Can't Found Trans Data: " + key[0] + "." + key[1])
        else:
            new_dict[key] = trans
            if key in missing_dict:
                missing_dict.pop(key)
    trans_dict = new_dict

    # print(trans_dict)

    # Create Trans Data
    print("Create Trans Data")
    for sheet_name in original_sheet_names:
        print("Modify Original Sheet: " + sheet_name)
        sheet = original_file[sheet_name]  # type: openpyxl.workbook.workbook.Worksheet
        # sheet.delete_cols(idx=3, amount=10)

        for i in range(sheet.max_row):
            cell = sheet.cell(row=i+1, column=1)
            if cell.value is not None:
                if (sheet_name, cell.value) in trans_dict:
                    sheet.cell(row=i+1, column=2, value=trans_dict[(sheet_name, cell.value)])

    print("Save Trans Data..")
    original_file.save('data\\T_Locales.xlsx')

    print("Begin Create Missing Data List File")
    missing_workbook = openpyxl.Workbook()
    sheet = missing_workbook["Sheet"]
    pos_row = 1
    for key in missing_dict:
        sheet.cell(row=pos_row, column=1, value=key[0] + "->" + key[1])
        sheet.cell(row=pos_row, column=2, value=missing_dict[key])
        pos_row += 1
    print("Save Missing Trans Data")
    missing_workbook.save('data\\Missing.xlsx')


